"""
The Solar Financial model gives users the expected financial output of a PV system based
on its costs and the amount energy it will likely produce. The model uses pvWatts, an
NLR model, to calculate how much energy the solar system will produce and the user’s
assumptions about the price of installing and maintaining the array.
"""

import json, shutil, datetime as dt
from os.path import join as pJoin
from numpy_financial import irr, npv
from openpyxl import Workbook
from pathlib import Path
import numpy as np
from math import prod

# OMF imports
from omf import weather
from omf.models import __neoMetaModel__
from omf.models.__neoMetaModel__ import *
from omf.solvers import pysam

# Model metadata:
modelName, template = __neoMetaModel__.metadata(__file__)
tooltip = "The solarFinancial model gives users the expected financial output of a PV system based on its costs and the amount energy it will likely produce."
hidden = False

def work(modelDir, inputDict):
	''' Run the model in its directory. '''
	derates = [
    float(inputDict.get("pvModuleDerate", 100)) / 100,
    float(inputDict.get("mismatch", 98)) / 100,
    float(inputDict.get("diodes", 99.5)) / 100,
    float(inputDict.get("dcWiring", 98)) / 100,
    float(inputDict.get("acWiring", 99)) / 100,
    float(inputDict.get("soiling", 95)) / 100,
    float(inputDict.get("shading", 100)) / 100,
    float(inputDict.get("sysAvail", 100)) / 100,
    float(inputDict.get("age", 100)) / 100
	]
	total_efficiency = prod(derates)
	total_loss_percent = (1 - total_efficiency) * 100
	inputDict["losses"] = total_loss_percent
	simStartDate = inputDict["simStartDate"]
	start = pd.to_datetime(inputDict["simStartDate"])
	# lat/long get checked in nlr_get_nsrdb_data
	lat = float( inputDict['latitude'] )
	long = float( inputDict['longitude'] )
	# parameter validation
	sys_design = pysam._pysam_sysDesignSetup(inputDict, lat, long)
	# We need DNI, DHI, GHI, windspeed, and temp
	attributes = ['dni,dhi,ghi,wind_speed,air_temperature']
	nrlAPIResponse = weather.nlr_get_nsrdb_data(data_set="goes_tmy", longitude=long, latitude=lat, year="tmy", api_key="rnvNJxNENljf60SBKGxkGVwkXls4IAKs1M8uZl56", attributes=attributes, filename=Path(modelDir,"output_tmy_data.csv"))
	requestSuccess = True if nrlAPIResponse.status_code == 200 else False
	# If getting the data was successful:
	# - Combine data + system parameters into pvwatts model and execute
	if requestSuccess:
		setYear = simStartDate[0:4]
		pvwatts_model, results_df = pysam.run_pvwatts(modelDir, sys_design=sys_design, dataFile="output_tmy_data.csv", setYear=True, year=setYear)
	else:
		raise Exception("solarFinancial.py: API request failed")
	outData = {}
	# Geodata output.
	outData['latitude'] = pvwatts_model.Outputs.lat
	outData['longitude'] = pvwatts_model.Outputs.lon
	outData['elev'] = pvwatts_model.Outputs.elev
	thirty_minute_start = pd.to_timedelta( 30, unit="minute") # The NSRDB data starts at 00:30 so everything is off by 30 minutes - that's what this is for.
	start = start + thirty_minute_start
	time_passed = pd.to_timedelta( int(inputDict['simLength']), unit=inputDict['simLengthUnits'])
	end = start + time_passed
	sim_df = results_df.loc[start:end]
	simLengthUnits = inputDict['simLengthUnits']
	if simLengthUnits == "minutes":
			freq = "t"
	elif simLengthUnits == "hours":
			freq = "h"
	elif simLengthUnits == "days":
			freq = "d"
	else:
			raise Exception()
	agg_df = sim_df.resample(freq).sum(numeric_only=True)
	simLengthUnits = inputDict.get("simLengthUnits","")
	startDateTime = simStartDate + " 00:00:00 UTC"
	outData["timeStamps"] = [ datetime.datetime.strftime(
	datetime.datetime.strptime(startDateTime[0:19],"%Y-%m-%d %H:%M:%S") + 
	datetime.timedelta(**{simLengthUnits:x}),"%Y-%m-%d %H:%M:%S") + " UTC" for x in range(int(inputDict["simLength"]))]

	# Weather output.
	outData["climate"] = {}
	outData["climate"]["Plane of Array Irradiance (W/m^2)"] = agg_df["poa"].tolist() if "poa" in agg_df else []
	outData["climate"]["Global Horizonal Irradiance (W/m^2)"] = agg_df["gh"].tolist() if "gh" in agg_df else []
	outData["climate"]["Beam Normal Irradiance (W/m^2)"] = agg_df["dn"].tolist() if "dn" in agg_df else []
	outData["climate"]["Diffuse Irradiance (W/m^2)"] = agg_df["df"].tolist() if "df" in agg_df else []
	outData["climate"]["Ambient Temperature (F)"] = agg_df["tamb"].tolist() if "tamb" in agg_df else []
	outData["climate"]["Cell Temperature (F)"] = agg_df["tcell"].tolist() if "tcell" in agg_df else []
	outData["climate"]["Wind Speed (m/s)"] = agg_df["wspd"].tolist() if "wspd" in agg_df else []
	outData['powerOutputAc'] = agg_df["ac"].tolist() if "ac" in agg_df else []
	invSizeWatts = float(inputDict.get('inverterSize', 0)) * 1000
	outData['InvClipped'] = [x if x < invSizeWatts else invSizeWatts for x in outData['powerOutputAc']]
	try:
		outData['percentClipped'] = 100 * (1.0 - sum(outData['InvClipped']) / sum(outData['powerOutputAc']))
	except ZeroDivisionError:
		outData['percentClipped'] = 0.0
	# Cashflow outputs.
	lifeSpan = int(inputDict.get("lifeSpan",30))
	lifeYears = list(range(1, 1 + lifeSpan))
	retailCost = float(inputDict.get("retailCost",0.0))
	degradation = float(inputDict.get("degradation",0.5))/100
	installCost = float(inputDict.get("installCost",0.0))
	discountRate = float(inputDict.get("discountRate", 7))/100
	outData["oneYearGenerationWh"] = sum(outData["powerOutputAc"])
	outData["lifeGenerationDollars"] = [retailCost*(1.0/1000)*outData["oneYearGenerationWh"]*(1.0-(x*degradation)) for x in lifeYears]
	outData["lifeOmCosts"] = [-1.0*float(inputDict["omCost"]) for x in lifeYears]
	outData["lifePurchaseCosts"] = [-1.0 * installCost] + [0 for x in lifeYears[1:]]
	srec = inputDict.get("srecCashFlow", "").split(",")
	outData["srecCashFlow"] = list(map(float,srec)) + [0 for x in lifeYears[len(srec):]]
	outData["netCashFlow"] = [x+y+z+a for (x,y,z,a) in zip(outData["lifeGenerationDollars"], outData["lifeOmCosts"], outData["lifePurchaseCosts"], outData["srecCashFlow"])]
	outData["cumCashFlow"] = [x for x in _runningSum(outData["netCashFlow"])]
	outData["ROI"] = __neoMetaModel__.roundSig(sum(outData["netCashFlow"]), 3) / (-1*__neoMetaModel__.roundSig(sum(outData["lifeOmCosts"]), 3) + -1*__neoMetaModel__.roundSig(sum(outData["lifePurchaseCosts"], 3)))
	outData["NPV"] = __neoMetaModel__.roundSig(npv(discountRate, outData["netCashFlow"]), 3)
	outData["lifeGenerationWh"] = sum(outData["powerOutputAc"])*lifeSpan	
	outData["lifeEnergySales"] = sum(outData["lifeGenerationDollars"])
	try:
		# The IRR function is very bad.
		outData["IRR"] = __neoMetaModel__.roundSig(irr(outData["netCashFlow"]), 3)
	except:
		outData["IRR"] = "Undefined"
	# Monthly aggregation outputs.
	months = {"Jan":0,"Feb":1,"Mar":2,"Apr":3,"May":4,"Jun":5,"Jul":6,"Aug":7,"Sep":8,"Oct":9,"Nov":10,"Dec":11}
	totMonNum = lambda x:sum([z for (y,z) in zip(outData["timeStamps"], outData["powerOutputAc"]) if y.startswith(simStartDate[0:4] + "-{0:02d}".format(x+1))])
	outData["monthlyGeneration"] = [[a, totMonNum(b)] for (a,b) in sorted(months.items(), key=lambda x:x[1])]
	# Heatmaped hour+month outputs.
	hours = list(range(24))
	from calendar import monthrange
	totHourMon = lambda h,m:sum([z for (y,z) in zip(outData["timeStamps"], outData["powerOutputAc"]) if y[5:7]=="{0:02d}".format(m+1) and y[11:13]=="{0:02d}".format(h+1)])
	outData["seasonalPerformance"] = [[x,y,totHourMon(x,y) / monthrange(int(simStartDate[:4]), y+1)[1]] for x in hours for y in months.values()]
	# Stdout/stderr.
	outData["stdout"] = "Success"
	outData["stderr"] = ""
	return outData

def _dumpDataToExcel(modelDir):
	""" Dump data into .xlsx file in model workspace """
	# TODO: Think about a universal function
	wb = Workbook()
	sh1 = wb.active
	sh1.title = "All Input Data"
	with open(pJoin(modelDir, "allInputData.json")) as f:
		inJson = json.load(f)
	size = len(inJson.keys())
	keys = list(inJson.keys())
	for i in range(size):
		sh1.cell(row=i + 1, column=1, value=keys[i])
	values = list(inJson.values())
	for i in range(size):
		sh1.cell(row=i + 1, column=2, value=values[i])
	with open(pJoin(modelDir, "allOutputData.json")) as f:
		outJson = json.load(f)
	sh1.cell(row=1, column=6, value="Lat")
	sh1.cell(row=1, column=7, value="Lon")
	sh1.cell(row=1, column=8, value="Elev")
	sh1.cell(row=2, column=6, value=outJson["lat"])
	sh1.cell(row=2, column=7, value=outJson["lon"])
	sh1.cell(row=2, column=8, value=outJson["elev"])

	sh2 = wb.create_sheet("Hourly Data")
	headers = [
		"TimeStamp",
		"Power(kW-AC)",
		"Power due to Inverter clipping(kW-AC)",
		"Plane of Array Irradiance (W/m^2)",
		"Global Horizontal Radiation(W/m^2)",
		"Wind Speed (m/s)",
		"Ambient Temperature (F)",
		"Cell Temperature (F)"
	]
	sh2.append(headers)

	for i in range(len(outJson["timeStamps"])):
		sh2.append([
			outJson["timeStamps"][i],
			outJson["powerOutputAc"][i],
			outJson["InvClipped"][i],
			outJson["climate"]["Plane of Array Irradiance (W/m^2)"][i],
			outJson["climate"]["Global Horizontal Radiation (W/m^2)"][i],
			outJson["climate"]["Wind Speed (m/s)"][i],
			outJson["climate"]["Ambient Temperature (F)"][i],
			outJson["climate"]["Cell Temperature (F)"][i]
		])

	sh2.freeze_panes = "B1"

	sh3 = wb.create_sheet("Monthly Data")
	sh3.cell(row=1, column=2, value="Monthly Generation")
	for i in range(24):
		sh3.cell(row=1, column=4 + i, value=i + 1)
	for i in range(12):
		sh3.cell(row=i + 2, column=1, value=outJson["monthlyGeneration"][i][0])
		sh3.cell(row=i + 2, column=2, value=outJson["monthlyGeneration"][i][1])
	for i in range(len(outJson["seasonalPerformance"])):
		sh3.cell(
			row=outJson["seasonalPerformance"][i][1] + 2,
			column=outJson["seasonalPerformance"][i][0] + 4,
			value=outJson["seasonalPerformance"][i][2]
		)
	sh3.freeze_panes = "D2"

	sh4 = wb.create_sheet("Annual Data")
	sh4.cell(row=1, column=1, value="Year No.")
	for i in range(len(outJson["netCashFlow"])):
		sh4.cell(row=i + 2, column=1, value=i)
	sh4.cell(row=1, column=2, value="Net Cash Flow ($)")
	sh4.cell(row=1, column=3, value="Life O&M Costs ($)")
	sh4.cell(row=1, column=4, value="Life Purchase Costs ($)")
	sh4.cell(row=1, column=5, value="Cumulative Cash Flow ($)")
	for i in range(len(outJson["netCashFlow"])):
		sh4.cell(row=i + 2, column=2, value=outJson["netCashFlow"][i])
		sh4.cell(row=i + 2, column=3, value=outJson["lifeOmCosts"][i])
		sh4.cell(row=i + 2, column=4, value=outJson["lifePurchaseCosts"][i])
		sh4.cell(row=i + 2, column=5, value=outJson["cumCashFlow"][i])
	sh4.cell(row=1, column=11, value="ROI")
	sh4.cell(row=2, column=11, value=outJson["ROI"])
	sh4.cell(row=1, column=12, value="NPV")
	sh4.cell(row=2, column=12, value=outJson["NPV"])
	sh4.cell(row=1, column=13, value="IRR")
	sh4.cell(row=2, column=13, value=outJson["IRR"])
	# sh4.cell(row=3, column=12, value="=NPV('All Input Data'!B15/100,'Annual Data'!B2:B31)")
	sh4.cell(row=3, column=13, value="=IRR(B2:B31)")
	filename = "omf.solarFinancial.xlsx"
	wb.save(pJoin(modelDir, filename))
	outJson["excel"] = filename
	with open(pJoin(modelDir,"allOutputData.json"),"w") as outFile:
		json.dump(outJson, outFile, indent=4)

def _runningSum(inList):
	''' Give a list of running sums of inList. '''
	return [sum(inList[:i+1]) for (i,val) in enumerate(inList)]

def new(modelDir):
	''' Create a new instance of this model. Returns true on success, false on failure. '''
	defaultInputs = {
		"simStartDate": "2013-01-01",
		"simLengthUnits": "hours",
		"modelType": modelName,
		"longitude": "-97.1292",
		"latitude": "33.2164",
		"simLength": "8760",
		"systemSize":"100",
		"installCost":"100000",
		"lifeSpan": "30",
		"degradation": "0.5",
		"retailCost": "0.10",
		"discountRate": "7",
		"pvModuleDerate": "100",
		"mismatch": "98",
		"diodes": "99.5",		
		"dcWiring": "98",
		"acWiring": "99",
		"soiling": "95",
		"shading": "100",
		"sysAvail": "100",
		"age": "100",		
		"inverterEfficiency": "92",
		"inverterSize": "75",
		"tilt": "45",
		"srecCashFlow": "5,5,3,3,2",
		"trackingMode":"0",
		"azimuth":"180",
		"runTime": "",
		"rotlim":"45.0",
		"gamma":"-0.45",
		"omCost": "1000"
	}
	return __neoMetaModel__.new(modelDir, defaultInputs)

@neoMetaModel_test_setup
def _tests():
	# Location
	"""
	Run this module's local smoke tests or debugging workflow.
	"""
	modelLoc = pJoin(__neoMetaModel__._omfDir,"data","Model","admin","Automated Testing of " + modelName)
	# Blow away old test results if necessary.
	try:
		shutil.rmtree(modelLoc)
	except:
		# No previous test results.
		pass
	# Create New.
	new(modelLoc)
	# Pre-run.
	__neoMetaModel__.renderAndShow(modelLoc)
	# Run the model.
	__neoMetaModel__.runForeground(modelLoc)
	# Show the output.
	__neoMetaModel__.renderAndShow(modelLoc)

if __name__ == '__main__':
	_tests()
